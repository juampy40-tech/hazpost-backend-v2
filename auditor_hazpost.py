"""
auditor_hazpost.py — Auditor web avanzado para hazpost.app
Genera auditoria_hazpost.xlsx (6 hojas) y auditoria_hazpost.txt

Nuevas capacidades v2:
  1. Analiza /dashboard (acceso público)
  2. Mide tiempos de carga reales (ms)
  3. Verifica enlaces rotos internos Y externos
  4. Evalúa rendimiento en móvil (simulado con UA móvil + checks HTML)
  5. Genera puntaje total /100 con recomendaciones priorizadas

Uso:  python3 auditor_hazpost.py
"""

import sys, time, difflib, re
from collections import defaultdict
from urllib.parse import urljoin, urlparse, urldefrag

# ── Auto-instalar openpyxl si falta ──────────────────────────────────────────
def ensure_packages():
    import importlib, subprocess
    try:
        importlib.import_module("openpyxl")
    except ImportError:
        print("[setup] instalando openpyxl...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])

ensure_packages()

import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Configuración ─────────────────────────────────────────────────────────────
BASE_URL   = "https://hazpost.app"
SEED_URLS  = [                          # Punto 1: incluir /dashboard
    "https://hazpost.app",
    "https://hazpost.app/dashboard",
    "https://hazpost.app/pricing",
    "https://hazpost.app/login",
    "https://hazpost.app/register",
    "https://hazpost.app/about",
    "https://hazpost.app/blog",
    "https://hazpost.app/features",
]
MAX_DEPTH            = 3
REQUEST_DELAY        = 0.4
TIMEOUT              = 12
EXT_LINK_TIMEOUT     = 8
SIMILARITY_THRESHOLD = 0.80
MAX_EXT_LINKS        = 30    # límite de links externos a verificar
OUTPUT_XLSX          = "auditoria_hazpost.xlsx"
OUTPUT_TXT           = "auditoria_hazpost.txt"

DESKTOP_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
MOBILE_UA = (                           # Punto 4: UA móvil para simulación
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_0 like Mac OS X) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.0 Mobile/15E148 Safari/604.1"
)

# ── Sesiones HTTP ─────────────────────────────────────────────────────────────
session_desktop = requests.Session()
session_desktop.headers.update({"User-Agent": DESKTOP_UA, "Accept-Language": "es-CO,es;q=0.9"})

session_mobile = requests.Session()
session_mobile.headers.update({"User-Agent": MOBILE_UA, "Accept-Language": "es-CO,es;q=0.9"})

# ── Estructuras de datos ──────────────────────────────────────────────────────
crawled_pages  = []
skills_found   = []
fallas         = []
duplicados     = []
mobile_results = []   # resultados de checks móvil
ext_links_done = set()
score_details  = {}   # se llena en calculate_score()

visited_urls = set()
queued_urls  = set()


# ═════════════════════════════════════════════════════════════════════════════
# 1. CRAWLER con medición de tiempos reales
# ═════════════════════════════════════════════════════════════════════════════

def is_internal(url: str) -> bool:
    p = urlparse(url); b = urlparse(BASE_URL)
    return p.netloc == "" or p.netloc == b.netloc

def normalize_url(url: str, base: str) -> str:
    full, _ = urldefrag(urljoin(base, url))
    return full.rstrip("/") or full

def should_crawl(url: str) -> bool:
    skip = (".png",".jpg",".jpeg",".gif",".svg",".webp",".pdf",".zip",
            ".mp4",".mp3",".woff",".woff2",".ttf",".eot",".ico",
            ".css",".js",".json",".xml",".txt",".map")
    p = urlparse(url)
    if not p.scheme.startswith("http"): return False
    if any(p.path.lower().endswith(e) for e in skip): return False
    if not is_internal(url): return False
    return True

def fetch_timed(url: str, session=None, method="get") -> tuple:
    """Retorna (response | None, load_time_ms)."""
    s = session or session_desktop
    t0 = time.time()
    try:
        fn = s.get if method == "get" else s.head
        r = fn(url, timeout=TIMEOUT, allow_redirects=True)
        ms = round((time.time() - t0) * 1000)
        return r, ms
    except requests.exceptions.Timeout:
        return None, TIMEOUT * 1000
    except Exception:
        return None, -1

def extract_page_data(url: str, depth: int, response, load_time_ms: int) -> dict:
    status = response.status_code if response else 0
    ct     = response.headers.get("Content-Type", "") if response else ""

    data = {
        "url": url, "depth": depth, "status": status,
        "load_time_ms": load_time_ms,
        "title": "", "h1": "", "meta_desc": "", "og_title": "", "og_desc": "",
        "has_viewport": False, "has_https": url.startswith("https"),
        "canonical": "", "robots_meta": "",
        "links": [], "ext_links": [],
        "images": [], "images_no_alt": 0,
        "scripts": [], "stylesheets": [],
        "body_text": "", "nav_items": [], "feature_cards": [], "cta_texts": [],
    }

    if not response or status != 200 or "text/html" not in ct:
        return data

    try:
        soup = BeautifulSoup(response.text, "lxml")
    except Exception:
        soup = BeautifulSoup(response.text, "html.parser")

    tt = soup.find("title")
    data["title"] = tt.get_text(strip=True) if tt else ""

    h1 = soup.find("h1")
    data["h1"] = h1.get_text(strip=True) if h1 else ""

    md = soup.find("meta", attrs={"name": re.compile(r"^description$", re.I)})
    data["meta_desc"] = (md.get("content", "") if md else "").strip()

    og = soup.find("meta", attrs={"property": "og:title"})
    data["og_title"] = (og.get("content", "") if og else "").strip()

    ogd = soup.find("meta", attrs={"property": "og:description"})
    data["og_desc"] = (ogd.get("content", "") if ogd else "").strip()

    vp = soup.find("meta", attrs={"name": re.compile(r"viewport", re.I)})
    data["has_viewport"] = vp is not None

    can = soup.find("link", rel=lambda r: r and "canonical" in r)
    data["canonical"] = can.get("href", "") if can else ""

    rob = soup.find("meta", attrs={"name": re.compile(r"robots", re.I)})
    data["robots_meta"] = (rob.get("content", "") if rob else "").strip()

    # Links internos y externos
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if not href or href.startswith(("#","javascript","mailto","tel")):
            continue
        full = normalize_url(href, url)
        if is_internal(full):
            data["links"].append(full)
        elif full.startswith("http"):
            data["ext_links"].append(full)

    # Imágenes (con count de las que no tienen alt)
    for img in soup.find_all("img"):
        src = normalize_url(img.get("src",""), url)
        if src: data["images"].append(src)
        if not img.get("alt","").strip():
            data["images_no_alt"] += 1

    for sc in soup.find_all("script", src=True):
        s2 = normalize_url(sc["src"], url)
        if s2: data["scripts"].append(s2)

    for lk in soup.find_all("link", rel=True):
        if "stylesheet" in (lk.get("rel") or []):
            h2 = normalize_url(lk.get("href",""), url)
            if h2: data["stylesheets"].append(h2)

    body = soup.find("body")
    if body:
        data["body_text"] = " ".join(body.get_text(separator=" ").split())[:3000]

    for nav in soup.find_all("nav"):
        for item in nav.find_all(["a","button","li"]):
            t = item.get_text(strip=True)
            if t and 1 < len(t) < 80:
                data["nav_items"].append(t)

    card_sel = [
        ("div", {"class": re.compile(r"card|feature|module|skill|plan|benefit|item", re.I)}),
        ("section", {"class": re.compile(r"feature|about|service|product", re.I)}),
        ("article", {}),
    ]
    seen_c = set()
    for tag, attrs in card_sel:
        for el in soup.find_all(tag, attrs):
            hd = el.find(["h2","h3","h4","strong","b"])
            if hd:
                t = hd.get_text(strip=True)
                if t and t not in seen_c and len(t) > 2:
                    seen_c.add(t)
                    dp = el.find("p")
                    data["feature_cards"].append({"nombre": t, "desc": (dp.get_text(strip=True)[:200] if dp else "")})

    cta_re = re.compile(
        r"(empieza|comienza|prueba|regístrate|crear|publicar|genera|descarga|"
        r"start|sign up|get started|try|create|publish|generate|free|gratis|"
        r"demo|solicitar|unirse|join)", re.I)
    for btn in soup.find_all(["button","a"], class_=re.compile(r"btn|cta|button", re.I)):
        t = btn.get_text(strip=True)
        if t and cta_re.search(t) and len(t) < 100:
            data["cta_texts"].append(t)

    return data


def crawl():
    # Sembrar URLs iniciales (incluye /dashboard)
    for seed in SEED_URLS:
        norm = normalize_url(seed, seed)
        if norm not in queued_urls:
            queued_urls.add(norm)

    queue = [(s, 0) for s in SEED_URLS]

    print(f"\n{'='*60}")
    print(f"  HazPost Web Auditor v2 — {BASE_URL}")
    print(f"  Seeds: {len(SEED_URLS)} URLs | Profundidad máx: {MAX_DEPTH}")
    print(f"{'='*60}\n")

    while queue:
        url, depth = queue.pop(0)
        if url in visited_urls:
            continue
        visited_urls.add(url)

        resp, ms = fetch_timed(url)
        time.sleep(REQUEST_DELAY)

        status = resp.status_code if resp else 0
        status_icon = "✓" if status == 200 else f"✗{status}"
        print(f"[{len(crawled_pages)+1:3d}] {status_icon} {ms:5d}ms  d={depth}  {url}")

        page_data = extract_page_data(url, depth, resp, ms)
        crawled_pages.append(page_data)

        if depth < MAX_DEPTH and resp and resp.status_code == 200:
            for link in page_data["links"]:
                norm = normalize_url(link, url)
                if norm not in queued_urls and should_crawl(norm):
                    queued_urls.add(norm)
                    queue.append((norm, depth + 1))

    print(f"\n✓ Rastreo completado: {len(crawled_pages)} páginas\n")


# ═════════════════════════════════════════════════════════════════════════════
# 2. EXTRACTOR DE SKILLS
# ═════════════════════════════════════════════════════════════════════════════

KNOWN_FEATURES = [
    ("Instagram","red_social"),("TikTok","red_social"),("Facebook","red_social"),
    ("WhatsApp","red_social"),("LinkedIn","red_social"),("Twitter","red_social"),
    ("Publicar","accion"),("Programar","accion"),("Calendario","modulo"),
    ("Generación de contenido","ia"),("Inteligencia artificial","ia"),
    ("IA","ia"),("AI","ia"),("GPT","ia"),("DALL-E","ia"),
    ("Imagen","contenido"),("Video","contenido"),("Reel","contenido"),
    ("Carrusel","contenido"),("Story","contenido"),("Historia","contenido"),
    ("Aprobación","flujo"),("Workflow","flujo"),("Cola","flujo"),
    ("Créditos","plan"),("Plan","plan"),("Suscripción","plan"),
    ("Agencia","plan"),("Multi-negocio","multitenancy"),
    ("Analytics","analítica"),("Estadísticas","analítica"),("Métricas","analítica"),
    ("Hashtag","seo"),("SEO","seo"),("Caption","contenido"),
    ("Marca","branding"),("Logo","branding"),("Tipografía","branding"),
    ("Fondo","branding"),("Paleta","branding"),
    ("Bulk","batch"),("Masivo","batch"),("Lote","batch"),
    ("Dashboard","modulo"),("Panel","modulo"),("Admin","admin"),
    ("Pago","pagos"),("Factura","pagos"),("Wompi","pagos"),
    ("Notificación","notif"),("Telegram","notif"),("Email","notif"),
    ("OAuth","auth"),("Login","auth"),("Registro","auth"),("Cuenta","auth"),
]

def extract_skills():
    print("[skills] extrayendo funcionalidades...")
    seen = set()
    for page in crawled_pages:
        full_text = f"{page['title']} {page['h1']} {page['body_text']}"
        for card in page["feature_cards"]:
            k = (card["nombre"].lower(), page["url"])
            if k not in seen:
                seen.add(k)
                skills_found.append({"Nombre": card["nombre"], "Página Origen": page["url"],
                    "Descripción": card["desc"] or "(sin descripción)", "Tipo": "feature_card", "Confianza": "alta"})
        for nav_t in set(page["nav_items"]):
            k = (nav_t.lower(), "nav")
            if k not in seen and len(nav_t) > 2:
                seen.add(k)
                skills_found.append({"Nombre": nav_t, "Página Origen": page["url"],
                    "Descripción": "Ítem de navegación", "Tipo": "nav_item", "Confianza": "media"})
        for kw, tipo in KNOWN_FEATURES:
            if re.search(r"\b" + re.escape(kw) + r"\b", full_text, re.I):
                k = (kw.lower(), tipo)
                if k not in seen:
                    seen.add(k)
                    skills_found.append({"Nombre": kw, "Página Origen": page["url"],
                        "Descripción": "Detectado en texto de página", "Tipo": tipo, "Confianza": "media"})
        for cta in set(page["cta_texts"]):
            k = (cta.lower(), "cta")
            if k not in seen:
                seen.add(k)
                skills_found.append({"Nombre": cta, "Página Origen": page["url"],
                    "Descripción": "Call-to-action detectado", "Tipo": "cta", "Confianza": "alta"})
    print(f"  → {len(skills_found)} items detectados")


# ═════════════════════════════════════════════════════════════════════════════
# 3. DETECTOR DE FALLAS — internos Y externos
# ═════════════════════════════════════════════════════════════════════════════

def check_resource(url: str, pagina_origen: str, tipo: str, timeout=8):
    try:
        r, ms = fetch_timed(url, method="head")
        if r is None:
            fallas.append({"URL": url, "Tipo de Falla": f"Timeout ({tipo})",
                "Detalle": f"No respondió en {timeout}s", "Página Origen": pagina_origen, "Tiempo (ms)": ms})
        elif r.status_code >= 400:
            fallas.append({"URL": url, "Tipo de Falla": f"HTTP {r.status_code} ({tipo})",
                "Detalle": f"Recurso retorna {r.status_code}", "Página Origen": pagina_origen, "Tiempo (ms)": ms})
        elif len(r.history) >= 3:
            fallas.append({"URL": url, "Tipo de Falla": "Redirect en cadena",
                "Detalle": f"{len(r.history)} redirects", "Página Origen": pagina_origen, "Tiempo (ms)": ms})
    except Exception as e:
        fallas.append({"URL": url, "Tipo de Falla": f"Error ({tipo})",
            "Detalle": str(e)[:200], "Página Origen": pagina_origen, "Tiempo (ms)": -1})
    time.sleep(0.2)


def detect_failures():
    print("[fallas] verificando páginas y recursos...")
    checked = set()
    ext_count = 0

    for page in crawled_pages:
        url = page["url"]

        # Status HTTP de la propia página
        if page["status"] == 0:
            fallas.append({"URL": url, "Tipo de Falla": "Inaccesible",
                "Detalle": "No se pudo conectar", "Página Origen": url, "Tiempo (ms)": page["load_time_ms"]})
        elif page["status"] >= 400:
            fallas.append({"URL": url, "Tipo de Falla": f"HTTP {page['status']}",
                "Detalle": f"Página retorna {page['status']}", "Página Origen": url, "Tiempo (ms)": page["load_time_ms"]})

        if page["status"] != 200:
            continue

        # Tiempo de carga lento (> 3 seg)
        if page["load_time_ms"] > 3000:
            fallas.append({"URL": url, "Tipo de Falla": "Carga lenta",
                "Detalle": f"Tardó {page['load_time_ms']}ms (umbral: 3000ms)",
                "Página Origen": url, "Tiempo (ms)": page["load_time_ms"]})

        # Meta tags
        if not page["title"]:
            fallas.append({"URL": url, "Tipo de Falla": "Meta faltante", "Detalle": "Sin <title>",
                "Página Origen": url, "Tiempo (ms)": 0})
        if not page["meta_desc"]:
            fallas.append({"URL": url, "Tipo de Falla": "Meta faltante", "Detalle": "Sin <meta description>",
                "Página Origen": url, "Tiempo (ms)": 0})
        if not page["og_title"]:
            fallas.append({"URL": url, "Tipo de Falla": "OG tag faltante", "Detalle": "Sin og:title",
                "Página Origen": url, "Tiempo (ms)": 0})
        if not page["h1"]:
            fallas.append({"URL": url, "Tipo de Falla": "SEO sin H1", "Detalle": "Página sin <h1>",
                "Página Origen": url, "Tiempo (ms)": 0})
        if not page["has_viewport"]:
            fallas.append({"URL": url, "Tipo de Falla": "Sin viewport meta",
                "Detalle": "Falta <meta name='viewport'> — afecta renderizado móvil",
                "Página Origen": url, "Tiempo (ms)": 0})
        if page["images_no_alt"] > 0:
            fallas.append({"URL": url, "Tipo de Falla": "Imágenes sin alt",
                "Detalle": f"{page['images_no_alt']} imágenes sin atributo alt (accesibilidad/SEO)",
                "Página Origen": url, "Tiempo (ms)": 0})

        # Recursos internos: imágenes, scripts, CSS
        for res_list, tipo in [(page["images"],"imagen"),(page["scripts"],"script"),(page["stylesheets"],"css")]:
            for ru in res_list:
                if ru in checked or not ru.startswith("http"): continue
                if is_internal(ru):
                    checked.add(ru)
                    check_resource(ru, url, tipo)

        # Links internos no rastreados
        for lk in page["links"]:
            if lk in checked or lk in visited_urls: continue
            if is_internal(lk) and should_crawl(lk):
                checked.add(lk)
                check_resource(lk, url, "link interno")

        # Links externos — verificar hasta MAX_EXT_LINKS en total
        for el in page["ext_links"]:
            if el in ext_links_done or ext_count >= MAX_EXT_LINKS: break
            ext_links_done.add(el)
            ext_count += 1
            check_resource(el, url, "link externo", timeout=EXT_LINK_TIMEOUT)

    print(f"  → {len(fallas)} fallas detectadas")


# ═════════════════════════════════════════════════════════════════════════════
# 4. EVALUACIÓN MÓVIL (simulada)
# ═════════════════════════════════════════════════════════════════════════════

def evaluate_mobile():
    print("[móvil] evaluando rendimiento en móvil simulado...")

    for page in crawled_pages:
        if page["status"] != 200:
            continue

        url = page["url"]
        result = {
            "URL": url,
            "Viewport Meta": "✓ Presente" if page["has_viewport"] else "✗ Faltante",
            "HTTPS": "✓ Sí" if page["has_https"] else "✗ No",
            "Tiempo carga (ms)": page["load_time_ms"],
            "Velocidad móvil": "",
            "Tiempo carga móvil (ms)": 0,
            "Imágenes sin alt": page["images_no_alt"],
            "Robots meta": page["robots_meta"] or "(ninguno)",
            "Problemas detectados": [],
        }

        # Fetch con UA móvil para comparar tiempo
        resp_m, ms_m = fetch_timed(url, session=session_mobile)
        time.sleep(REQUEST_DELAY)
        result["Tiempo carga móvil (ms)"] = ms_m

        # Clasificar velocidad
        if ms_m <= 1500:
            result["Velocidad móvil"] = "Rápida (< 1.5s)"
        elif ms_m <= 3000:
            result["Velocidad móvil"] = "Aceptable (1.5–3s)"
        else:
            result["Velocidad móvil"] = "Lenta (> 3s)"

        problemas = []
        if not page["has_viewport"]:
            problemas.append("Sin viewport meta — el sitio no es responsive")
        if ms_m > 3000:
            problemas.append(f"Carga lenta en móvil: {ms_m}ms")
        if page["images_no_alt"] > 0:
            problemas.append(f"{page['images_no_alt']} imágenes sin alt")
        if not page["has_https"]:
            problemas.append("Sin HTTPS — inseguro en móvil")

        # Parsear HTML móvil y buscar problemas adicionales
        if resp_m and resp_m.status_code == 200 and "text/html" in resp_m.headers.get("Content-Type",""):
            try:
                soup_m = BeautifulSoup(resp_m.text, "lxml")
            except Exception:
                soup_m = BeautifulSoup(resp_m.text, "html.parser")

            # Detectar fuentes muy pequeñas hardcoded en style inline
            small_font = re.findall(r'font-size\s*:\s*([0-9]+)px', resp_m.text, re.I)
            tiny = [int(f) for f in small_font if int(f) < 12]
            if tiny:
                problemas.append(f"Fuentes < 12px detectadas: {tiny[:5]}")

            # Verificar que no haya tablas usadas como layout (anti-patrón mobile)
            tables = soup_m.find_all("table")
            non_data_tables = [t for t in tables if not t.find(["th", "thead"])]
            if len(non_data_tables) > 2:
                problemas.append(f"{len(non_data_tables)} tablas usadas como layout (anti-patrón mobile)")

        result["Problemas detectados"] = "; ".join(problemas) if problemas else "Ninguno detectado"
        mobile_results.append(result)

    print(f"  → {len(mobile_results)} páginas evaluadas en móvil")


# ═════════════════════════════════════════════════════════════════════════════
# 5. DETECTOR DE DUPLICADOS
# ═════════════════════════════════════════════════════════════════════════════

def detect_duplicates():
    print("[duplicados] buscando contenido similar...")
    pages_ok = [p for p in crawled_pages if p["status"] == 200]
    compared = 0
    for i in range(len(pages_ok)):
        for j in range(i+1, len(pages_ok)):
            pa, pb = pages_ok[i], pages_ok[j]
            for campo, ta, tb in [
                ("Título", pa["title"], pb["title"]),
                ("Meta Description", pa["meta_desc"], pb["meta_desc"]),
                ("Contenido", pa["body_text"][:500], pb["body_text"][:500]),
            ]:
                if not ta or not tb: continue
                sim = difflib.SequenceMatcher(None, ta.lower(), tb.lower()).ratio()
                if sim >= SIMILARITY_THRESHOLD:
                    duplicados.append({
                        "URL A": pa["url"], "Texto A": ta[:200],
                        "URL B": pb["url"], "Texto B": tb[:200],
                        "Similitud": round(sim, 3), "Tipo": campo,
                        "Sugerencia": f"Consolidar en {pa['url']} o diferenciar el contenido",
                    })
            compared += 1
    print(f"  → {len(duplicados)} pares similares (de {compared} comparaciones)")


# ═════════════════════════════════════════════════════════════════════════════
# 6. GENERADOR DE PROPUESTAS
# ═════════════════════════════════════════════════════════════════════════════

PROPUESTAS_BASE = [
    {"Categoría":"Analytics avanzados","Prioridad":"Alta","Propuesta":"Panel con métricas reales de Instagram/TikTok/Facebook: alcance, engagement rate, mejor hora de publicación","Impacto":"Retención — los usuarios quieren ver ROI"},
    {"Categoría":"A/B Testing de contenido","Prioridad":"Media","Propuesta":"2 variantes de post con medición automática del ganador","Impacto":"Diferenciador de mercado"},
    {"Categoría":"Colaboración en equipo","Prioridad":"Alta","Propuesta":"Roles (diseñador, redactor, aprobador) con permisos por negocio","Impacto":"Escala el plan Agencia"},
    {"Categoría":"Programación best-time","Prioridad":"Alta","Propuesta":"Sugerir hora óptima de publicación por cuenta basándose en historial de engagement","Impacto":"Aumenta alcance orgánico sin esfuerzo"},
    {"Categoría":"Exportación reportes PDF","Prioridad":"Alta","Propuesta":"Reporte mensual PDF/PPT con logo del negocio para que agencias lo envíen a sus clientes","Impacto":"Cierre de plan Agencia"},
    {"Categoría":"Feed preview visual","Prioridad":"Media","Propuesta":"Vista previa del feed de Instagram con posts programados","Impacto":"Funcionalidad pedida por community managers"},
    {"Categoría":"Hashtag research","Prioridad":"Media","Propuesta":"Herramienta de investigación de hashtags con volumen y competencia","Impacto":"Mejora el alcance orgánico"},
    {"Categoría":"Integración CRM","Prioridad":"Media","Propuesta":"Sincronizar leads de redes con HubSpot, Pipedrive o Google Sheets","Impacto":"Atrae clientes B2B"},
    {"Categoría":"Chatbot de soporte","Prioridad":"Media","Propuesta":"Chatbot con IA dentro del dashboard para responder dudas de uso","Impacto":"Reduce tickets, mejora onboarding"},
    {"Categoría":"Plantillas por industria","Prioridad":"Alta","Propuesta":"Biblioteca de plantillas copy+imagen por sector para acelerar el onboarding","Impacto":"Reduce tiempo al primer post"},
    {"Categoría":"LinkedIn Publishing","Prioridad":"Media","Propuesta":"Soporte de publicación directa en LinkedIn","Impacto":"Abre mercado B2B"},
    {"Categoría":"Integración Canva","Prioridad":"Media","Propuesta":"Importar diseños de Canva como fondo de posts","Impacto":"Captura usuarios que ya usan Canva"},
    {"Categoría":"Reels música personalizada","Prioridad":"Alta","Propuesta":"Biblioteca de tracks libres de derechos con selector de segmento","Impacto":"Diferenciador clave TikTok/Reels"},
    {"Categoría":"Watermark de protección","Prioridad":"Baja","Propuesta":"Watermark automático con logo del negocio en imágenes generadas","Impacto":"Protección de contenido"},
    {"Categoría":"Revisión ortográfica/tono","Prioridad":"Media","Propuesta":"Corrector de estilo y tono en el editor de captions","Impacto":"Mejora calidad sin esfuerzo del usuario"},
]

def generate_proposals():
    print("[propuestas] generando recomendaciones...")
    all_text = " ".join(p["body_text"] for p in crawled_pages if p["status"]==200).lower()
    for prop in PROPUESTAS_BASE:
        kws = prop["Categoría"].lower().split()
        presente = any(kw in all_text for kw in kws if len(kw) > 4)
        prop["Estado Detectado"] = "Posiblemente presente" if presente else "No detectado"

    missing_meta = sum(1 for f in fallas if "meta" in f["Tipo de Falla"].lower())
    if missing_meta > 0:
        PROPUESTAS_BASE.insert(0, {"Categoría":"SEO técnico","Prioridad":"Alta",
            "Propuesta":f"Completar meta tags en {missing_meta} página(s): og:title, description, H1",
            "Impacto":"Mejora inmediata de indexación y compartibilidad","Estado Detectado":"Falla activa"})

    no_viewport = sum(1 for f in fallas if "viewport" in f["Tipo de Falla"].lower())
    if no_viewport > 0:
        PROPUESTAS_BASE.insert(1, {"Categoría":"Mobile — viewport","Prioridad":"Alta",
            "Propuesta":f"Agregar <meta name='viewport'> en {no_viewport} página(s)",
            "Impacto":"Renderizado correcto en todos los dispositivos móviles","Estado Detectado":"Falla activa"})

    slow = sum(1 for f in fallas if "Carga lenta" in f["Tipo de Falla"])
    if slow > 0:
        PROPUESTAS_BASE.insert(2, {"Categoría":"Performance","Prioridad":"Alta",
            "Propuesta":f"{slow} página(s) superan 3s de carga — optimizar assets, habilitar CDN, lazy loading",
            "Impacto":"Reducción de bounce rate, mejor ranking en Google","Estado Detectado":"Falla activa"})

    print(f"  → {len(PROPUESTAS_BASE)} propuestas generadas")


# ═════════════════════════════════════════════════════════════════════════════
# 7. PUNTAJE TOTAL /100
# ═════════════════════════════════════════════════════════════════════════════

def calculate_score() -> dict:
    print("[puntaje] calculando score total...")

    pages_ok = [p for p in crawled_pages if p["status"] == 200]
    total_pages = max(len(pages_ok), 1)

    # ── SEO (25 pts) ─────────────────────────────────────────────────────────
    seo = 25
    no_title    = sum(1 for p in pages_ok if not p["title"])
    no_desc     = sum(1 for p in pages_ok if not p["meta_desc"])
    no_og       = sum(1 for p in pages_ok if not p["og_title"])
    no_h1       = sum(1 for p in pages_ok if not p["h1"])
    seo -= min(8,  round(no_title / total_pages * 8))
    seo -= min(6,  round(no_desc  / total_pages * 6))
    seo -= min(5,  round(no_og    / total_pages * 5))
    seo -= min(6,  round(no_h1    / total_pages * 6))
    seo  = max(seo, 0)

    # ── Rendimiento (25 pts) ──────────────────────────────────────────────────
    perf = 25
    avg_ms = (sum(p["load_time_ms"] for p in pages_ok) / total_pages) if pages_ok else 0
    slow_pages = sum(1 for p in pages_ok if p["load_time_ms"] > 3000)
    if avg_ms > 5000:   perf -= 15
    elif avg_ms > 3000: perf -= 10
    elif avg_ms > 1500: perf -= 5
    perf -= min(10, slow_pages * 3)
    perf  = max(perf, 0)

    # ── Mobile (20 pts) ───────────────────────────────────────────────────────
    mob = 20
    no_vp   = sum(1 for p in pages_ok if not p["has_viewport"])
    no_https= sum(1 for p in pages_ok if not p["has_https"])
    mob -= min(12, round(no_vp    / total_pages * 12))
    mob -= min(5,  round(no_https / total_pages * 5))
    mob_slow = sum(1 for r in mobile_results if r["Tiempo carga móvil (ms)"] > 3000)
    mob -= min(3, mob_slow)
    mob  = max(mob, 0)

    # ── Accesibilidad (15 pts) ────────────────────────────────────────────────
    acc = 15
    total_no_alt = sum(p["images_no_alt"] for p in pages_ok)
    broken_links = sum(1 for f in fallas if "HTTP 4" in f["Tipo de Falla"])
    acc -= min(8, total_no_alt // 3)
    acc -= min(7, broken_links * 2)
    acc  = max(acc, 0)

    # ── Contenido / Estructura (15 pts) ───────────────────────────────────────
    cont = 15
    has_https_all = all(p["has_https"] for p in pages_ok)
    if not has_https_all: cont -= 5
    if len(pages_ok) < 3: cont -= 5    # sitio con muy pocas páginas rastreables
    if duplicados:         cont -= min(5, len(duplicados) * 2)
    cont = max(cont, 0)

    total = seo + perf + mob + acc + cont

    label = (
        "Excelente" if total >= 90 else
        "Muy bueno" if total >= 80 else
        "Bueno"     if total >= 70 else
        "Regular"   if total >= 55 else
        "Necesita mejora"
    )

    detalles = [
        {"Categoría": "SEO (títulos, metas, H1, OG tags)", "Puntos Obtenidos": seo, "Puntos Máximos": 25,
         "% Logrado": f"{seo/25*100:.0f}%", "Observación": f"title={no_title} sin título, {no_desc} sin desc, {no_h1} sin H1"},
        {"Categoría": "Rendimiento / Velocidad de carga", "Puntos Obtenidos": perf, "Puntos Máximos": 25,
         "% Logrado": f"{perf/25*100:.0f}%", "Observación": f"Promedio {avg_ms:.0f}ms, {slow_pages} pág. lentas"},
        {"Categoría": "Mobile / Responsive", "Puntos Obtenidos": mob, "Puntos Máximos": 20,
         "% Logrado": f"{mob/20*100:.0f}%", "Observación": f"{no_vp} sin viewport, {mob_slow} lentas en móvil"},
        {"Categoría": "Accesibilidad (alt texts, links rotos)", "Puntos Obtenidos": acc, "Puntos Máximos": 15,
         "% Logrado": f"{acc/15*100:.0f}%", "Observación": f"{total_no_alt} imágenes sin alt, {broken_links} links rotos"},
        {"Categoría": "Contenido / Estructura", "Puntos Obtenidos": cont, "Puntos Máximos": 15,
         "% Logrado": f"{cont/15*100:.0f}%", "Observación": f"HTTPS: {'Sí' if has_https_all else 'No'}, {len(duplicados)} duplicados"},
        {"Categoría": "── TOTAL ──", "Puntos Obtenidos": total, "Puntos Máximos": 100,
         "% Logrado": f"{total}%", "Observación": f"Calificación: {label}"},
    ]

    # Recomendaciones priorizadas (ordenadas por impacto/puntos perdidos)
    recs = []
    gaps = [
        (25-seo,  "SEO",           f"Completar meta tags, OG tags y H1 en todas las páginas"),
        (25-perf, "Performance",   f"Optimizar assets, habilitar CDN, reducir tiempo promedio ({avg_ms:.0f}ms)"),
        (20-mob,  "Mobile",        f"Agregar viewport meta en {no_vp} página(s), optimizar carga móvil"),
        (15-acc,  "Accesibilidad", f"Añadir atributo alt a {total_no_alt} imagen(es), corregir {broken_links} link(s) roto(s)"),
        (15-cont, "Contenido",     f"Activar HTTPS en todas las páginas, eliminar {len(duplicados)} duplicado(s)"),
    ]
    for pts_perdidos, cat, desc in sorted(gaps, key=lambda x: -x[0]):
        if pts_perdidos > 0:
            recs.append({"Prioridad": "Alta" if pts_perdidos >= 10 else "Media" if pts_perdidos >= 5 else "Baja",
                "Categoría": cat, "Puntos recuperables": pts_perdidos,
                "Acción recomendada": desc})

    score_details["total"] = total
    score_details["label"] = label
    score_details["avg_ms"] = avg_ms
    score_details["detalles"] = detalles
    score_details["recomendaciones"] = recs

    print(f"  → Puntaje total: {total}/100 — {label}")
    return score_details


# ═════════════════════════════════════════════════════════════════════════════
# 8. EXPORTADOR EXCEL + TXT
# ═════════════════════════════════════════════════════════════════════════════

HEADER_FILL  = PatternFill("solid", fgColor="1A1A2E")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
ALT_FILL     = PatternFill("solid", fgColor="F0F4FF")
SCORE_FILL   = PatternFill("solid", fgColor="16A34A")   # verde para el total
WARN_FILL    = PatternFill("solid", fgColor="DC2626")   # rojo para score bajo

def style_sheet(ws, score_row=None):
    for cell in ws[1]:
        cell.font = HEADER_FONT; cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for ri, row in enumerate(ws.iter_rows(min_row=2), 2):
        fill = ALT_FILL if ri % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for ci, col in enumerate(ws.columns, 1):
        mx = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(ci)].width = min(max(mx+4, 14), 80)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

def _write_df(ws, df):
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append([str(v) if v is not None else "" for v in row])

def export_excel():
    print("[excel] escribiendo auditoria_hazpost.xlsx ...")
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)

    # Hoja 1: Skills Activos
    df1 = pd.DataFrame(skills_found)[["Nombre","Tipo","Confianza","Página Origen","Descripción"]].drop_duplicates(subset=["Nombre","Tipo"]) if skills_found else pd.DataFrame(columns=["Nombre","Tipo","Confianza","Página Origen","Descripción"])
    ws1 = wb.create_sheet("Skills Activos"); _write_df(ws1, df1); style_sheet(ws1)

    # Hoja 2: Estructura Web
    struct = [{"URL": p["url"], "Profundidad": p["depth"], "Status": p["status"],
        "Carga desktop (ms)": p["load_time_ms"], "Velocidad": ("Rápida" if p["load_time_ms"]<=1500 else "Aceptable" if p["load_time_ms"]<=3000 else "Lenta"),
        "Título": p["title"] or "(sin título)", "H1": p["h1"] or "(sin H1)",
        "Meta Desc": (p["meta_desc"] or "(faltante)")[:120], "OG Title": p["og_title"] or "(faltante)",
        "Viewport": "✓" if p["has_viewport"] else "✗", "HTTPS": "✓" if p["has_https"] else "✗",
        "Imágenes sin alt": p["images_no_alt"], "Links internos": len(p["links"]), "Links externos": len(p["ext_links"])
    } for p in crawled_pages]
    ws2 = wb.create_sheet("Estructura Web"); _write_df(ws2, pd.DataFrame(struct)); style_sheet(ws2)

    # Hoja 3: Fallas y Problemas
    df3 = pd.DataFrame(fallas)[["URL","Tipo de Falla","Detalle","Página Origen","Tiempo (ms)"]] if fallas else pd.DataFrame(columns=["URL","Tipo de Falla","Detalle","Página Origen","Tiempo (ms)"])
    ws3 = wb.create_sheet("Fallas y Problemas"); _write_df(ws3, df3); style_sheet(ws3)

    # Hoja 4: Duplicados
    df4 = pd.DataFrame(duplicados)[["Tipo","Similitud","URL A","Texto A","URL B","Texto B","Sugerencia"]] if duplicados else pd.DataFrame(columns=["Tipo","Similitud","URL A","Texto A","URL B","Texto B","Sugerencia"])
    if not duplicados:
        df4.loc[0] = ["(ninguno)",0,BASE_URL,"",BASE_URL,"","Sin duplicados detectados"]
    ws4 = wb.create_sheet("Duplicados"); _write_df(ws4, df4); style_sheet(ws4)

    # Hoja 5: Propuestas de Mejora
    df5 = pd.DataFrame(PROPUESTAS_BASE)
    if "Estado Detectado" not in df5.columns: df5["Estado Detectado"] = "No evaluado"
    df5 = df5[["Categoría","Prioridad","Estado Detectado","Propuesta","Impacto"]]
    ws5 = wb.create_sheet("Propuestas de Mejora"); _write_df(ws5, df5); style_sheet(ws5)

    # Hoja 6: Puntaje Total /100
    sd = score_details
    df6a = pd.DataFrame(sd.get("detalles", []))[["Categoría","Puntos Obtenidos","Puntos Máximos","% Logrado","Observación"]]
    ws6 = wb.create_sheet("Puntaje Total")
    ws6.append(["AUDITORÍA HAZPOST.APP — PUNTAJE TOTAL"])
    ws6["A1"].font = Font(bold=True, size=14, color="1A1A2E")
    ws6.append([f"Score: {sd.get('total', 0)}/100 — {sd.get('label','')}", "", f"Tiempo promedio: {sd.get('avg_ms',0):.0f}ms"])
    ws6.append([])
    ws6.append(["DESGLOSE DE PUNTAJE"])
    ws6["A4"].font = Font(bold=True)
    _write_df(ws6, df6a)
    ws6.append([])
    ws6.append(["RECOMENDACIONES PRIORIZADAS"])
    ws6[f"A{ws6.max_row}"].font = Font(bold=True)
    df6b = pd.DataFrame(sd.get("recomendaciones", []))[["Prioridad","Categoría","Puntos recuperables","Acción recomendada"]]
    _write_df(ws6, df6b)
    style_sheet(ws6)

    # Añadir móvil a Estructura Web como sub-info
    if mobile_results:
        ws_mob = wb.create_sheet("Rendimiento Móvil")
        df_mob = pd.DataFrame(mobile_results)[["URL","Viewport Meta","HTTPS","Tiempo carga (ms)","Tiempo carga móvil (ms)","Velocidad móvil","Imágenes sin alt","Problemas detectados"]]
        _write_df(ws_mob, df_mob); style_sheet(ws_mob)

    wb.save(OUTPUT_XLSX)
    print(f"  ✓ {OUTPUT_XLSX} guardado ({wb.sheetnames})")


def export_txt():
    print("[txt] escribiendo auditoria_hazpost.txt ...")
    sd = score_details
    lines = []
    SEP = "=" * 70

    lines += [SEP, "  AUDITORÍA WEB — hazpost.app  (v2)", f"  {time.strftime('%d/%m/%Y %H:%M:%S')}",
              f"  Páginas rastreadas: {len(crawled_pages)} | Seeds: {len(SEED_URLS)}", SEP, ""]

    # Puntaje destacado al inicio
    total = sd.get("total", 0)
    bar = "█" * (total // 5) + "░" * (20 - total // 5)
    lines += [f"  PUNTAJE TOTAL: {total}/100  [{bar}]  {sd.get('label','')}",
              f"  Promedio carga: {sd.get('avg_ms',0):.0f}ms", ""]
    for d in sd.get("detalles", []):
        pct = int(d["Puntos Obtenidos"]) / int(d["Puntos Máximos"]) * 100 if d["Puntos Máximos"] != "100" else int(d["Puntos Obtenidos"])
        lines.append(f"  {d['Categoría']:45s} {d['Puntos Obtenidos']:>3}/{d['Puntos Máximos']}  {d['Observación']}")
    lines.append("")

    # 1. Skills
    lines += ["1. SKILLS / FUNCIONALIDADES DETECTADAS", "-"*50]
    unique_skills = list({s["Nombre"]: s for s in skills_found}.values())
    for s in unique_skills[:40]:
        lines.append(f"  [{s['Tipo'].upper():15s}] {s['Nombre']}")
    lines += [f"\n  Total: {len(skills_found)} items", ""]

    # 2. Estructura + tiempos
    lines += ["2. ESTRUCTURA WEB Y TIEMPOS DE CARGA", "-"*50]
    for p in crawled_pages:
        icon = "✓" if p["status"] == 200 else "✗"
        spd  = "🟢" if p["load_time_ms"]<=1500 else "🟡" if p["load_time_ms"]<=3000 else "🔴"
        lines.append(f"  {icon} {spd} {p['load_time_ms']:5d}ms  [{p['depth']}]  {p['url']}  ({p['status']})")
        if p["title"]: lines.append(f"       {p['title'][:80]}")
    lines.append("")

    # 3. Fallas
    lines += ["3. FALLAS Y PROBLEMAS", "-"*50]
    tipo_counts = defaultdict(int)
    for f in fallas: tipo_counts[f["Tipo de Falla"]] += 1
    for t, c in sorted(tipo_counts.items(), key=lambda x: -x[1]):
        lines.append(f"  {c:3d}x  {t}")
    lines.append("")
    for f in fallas[:25]:
        lines.append(f"  ✗ {f['Tipo de Falla']}: {f['URL'][:75]}")
        lines.append(f"    → {f['Detalle'][:120]}")
    if len(fallas) > 25: lines.append(f"  ... y {len(fallas)-25} más (ver Excel)")
    lines += [f"\n  Total: {len(fallas)} fallas", ""]

    # 4. Duplicados
    lines += ["4. CONTENIDO DUPLICADO (umbral ≥ 80%)", "-"*50]
    if duplicados:
        for d in duplicados[:10]:
            lines += [f"  [{d['Tipo']}] Similitud: {d['Similitud']*100:.0f}%",
                      f"    A: {d['URL A']}", f"    B: {d['URL B']}",
                      f"    → {d['Sugerencia'][:100]}", ""]
    else: lines.append("  (Sin duplicados detectados)")
    lines += [f"  Total: {len(duplicados)} pares", ""]

    # 5. Propuestas
    lines += ["5. PROPUESTAS DE MEJORA (priorizadas)", "-"*50]
    for i, prop in enumerate(PROPUESTAS_BASE, 1):
        lines += [f"  {i:2d}. [{prop['Prioridad'].upper():5s}] {prop['Categoría']}",
                  f"       {prop['Propuesta'][:120]}",
                  f"       Impacto: {prop['Impacto'][:100]}", ""]

    # 6. Recomendaciones del score
    lines += ["6. RECOMENDACIONES PRIORIZADAS POR PUNTAJE", "-"*50]
    for r in sd.get("recomendaciones", []):
        lines += [f"  [{r['Prioridad'].upper():5s}] {r['Categoría']} — +{r['Puntos recuperables']} pts posibles",
                  f"       {r['Acción recomendada'][:130]}", ""]

    lines += [SEP, "  FIN DEL INFORME", SEP]

    with open(OUTPUT_TXT, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  ✓ {OUTPUT_TXT} guardado")


# ═════════════════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════════════════

def main():
    t0 = time.time()
    crawl()
    extract_skills()
    detect_failures()
    evaluate_mobile()
    detect_duplicates()
    generate_proposals()
    calculate_score()
    export_excel()
    export_txt()
    elapsed = time.time() - t0
    sd = score_details
    print(f"\n{'='*60}")
    print(f"  ✅ Auditoría v2 completa en {elapsed:.1f}s")
    print(f"  🎯 Puntaje: {sd.get('total',0)}/100 — {sd.get('label','')}")
    print(f"  📊 {OUTPUT_XLSX}")
    print(f"  📄 {OUTPUT_TXT}")
    print(f"{'='*60}\n")

if __name__ == "__main__":
    main()
